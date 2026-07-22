import pyautogui
import pyperclip
import subprocess
import time
import os
from datetime import datetime
from openpyxl import Workbook

# ----------------------------------------
# Configuration
# ----------------------------------------

pyautogui.PAUSE = 1
pyautogui.FAILSAFE = True

# ----------------------------------------
# Current Date & Time
# ----------------------------------------

now = datetime.now()

current_datetime = now.strftime("%Y-%m-%d %H:%M:%S")
current_date = now.strftime("%Y-%m-%d")

file_name = f"daily_report_{current_date}.xlsx"

project_path = os.getcwd()

file_path = os.path.join(project_path, file_name)

# ----------------------------------------
# Step 1 - Open Chrome
# ----------------------------------------

print("Opening Chrome...")

subprocess.Popen([
    "open",
    "-a",
    "Google Chrome",
    "https://www.cricbuzz.com/cricket-match/live-scores"
])

time.sleep(8)

# ----------------------------------------
# Step 2 - Copy Live Score
# ----------------------------------------

# Adjust Tab count if required
pyautogui.press("tab", presses=17, interval=0.3)

pyautogui.hotkey("command", "a" , interval=0.3)
pyautogui.hotkey("command", "c" , interval=0.3)

time.sleep(2)

score = pyperclip.paste()

if score.strip() == "":
    score = "Live Cricket Score"

print(score)

# ----------------------------------------
# Step 3 - Create Excel Workbook
# ----------------------------------------

print("Creating Excel file...")

wb = Workbook()

ws = wb.active

ws.title = "Daily Report"

ws["A1"] = "Date & Time"
ws["B1"] = "Match Details"
ws["C1"] = "Comment"

ws["A2"] = current_datetime
ws["B2"] = score
ws["C2"] = "Daily Cricket Update"

wb.save(file_path)

print("Excel Saved:", file_path)

# ----------------------------------------
# Step 4 - Open Excel
# ----------------------------------------

print("Opening Excel...")

subprocess.Popen([
    "open",
    file_path
])

time.sleep(8)

# ----------------------------------------
# Step 5 - Screenshot
# ----------------------------------------

screenshot = os.path.join(
    project_path,
    f"daily_report_{current_date}.png"
)

pyautogui.screenshot(screenshot)

print("Screenshot Saved:", screenshot)

print("Automation Completed Successfully.")